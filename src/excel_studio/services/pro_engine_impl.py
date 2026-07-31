from __future__ import annotations

import contextlib
import copy
import json
import re
import unicodedata
from collections import OrderedDict
from collections.abc import Callable, Iterable
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Protocol

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from excel_studio.models.advanced import (
    AdvancedMergeTaskConfig,
    DuplicateKeyPolicy,
    JoinConflictPolicy,
)
from excel_studio.models.operations import (
    AggregateMethod,
    BlankValuePolicy,
    DedupeMode,
    ExistingFilePolicy,
    FieldStrategy,
    MergeMode,
    MergeTaskConfig,
    OutputFormat,
    OutputMode,
    OutputOptions,
    SheetConflictPolicy,
    SplitDistribution,
    SplitMode,
    SplitTaskConfig,
)
from excel_studio.models.result import ReconciliationResult, TaskResult
from excel_studio.services.reconciliation import reconcile_rows
from excel_studio.services.table_reader import read_sheet_frame
from excel_studio.utils.naming import (
    render_naming_template,
    sanitize_filename,
    sanitize_sheet_name,
    unique_output_path,
)

EXCEL_MAX_ROWS = 1_048_576
EXCEL_MAX_COLUMNS = 16_384


class TaskControlProtocol(Protocol):
    def checkpoint(self) -> None: ...


@dataclass(slots=True)
class EngineHooks:
    control: TaskControlProtocol | None = None
    log: Callable[[str], None] | None = None
    progress: Callable[[int, str, str], None] | None = None

    def checkpoint(self) -> None:
        if self.control is not None:
            self.control.checkpoint()

    def write_log(self, message: str) -> None:
        if self.log is not None:
            self.log(message)

    def update(self, percent: int, step: str, current_file: str = "") -> None:
        if self.progress is not None:
            self.progress(percent, step, current_file)


def _sheet_names(path: Path, configured: dict[Path, list[str]]) -> list[str]:
    explicit = configured.get(path)
    if explicit:
        return explicit
    extension = path.suffix.casefold()
    if extension in {".csv", ".tsv"}:
        return [path.stem]
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(
            path,
            read_only=True,
            keep_vba=extension == ".xlsm",
            keep_links=False,
        )
        try:
            return [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
        finally:
            workbook.close()
    return list(pd.ExcelFile(path).sheet_names)


def _resolve_output(directory: Path, stem: str, suffix: str, options: OutputOptions) -> Path | None:
    directory.mkdir(parents=True, exist_ok=True)
    requested = directory / f"{sanitize_filename(stem)}{suffix}"
    policy = options.existing_file_policy
    if options.overwrite:
        policy = ExistingFilePolicy.OVERWRITE
    if not requested.exists() or policy == ExistingFilePolicy.OVERWRITE:
        return requested
    if policy == ExistingFilePolicy.SKIP:
        return None
    return unique_output_path(directory, sanitize_filename(stem), suffix)


def _excel_safe_value(value: Any) -> Any:
    if value is None or (not isinstance(value, (list, dict, tuple, set)) and pd.isna(value)):
        return None
    if hasattr(value, "to_pydatetime"):
        try:
            return value.to_pydatetime()
        except Exception:
            pass
    if isinstance(value, (list, dict, tuple, set)):
        return json.dumps(value, ensure_ascii=False, default=str)
    return value


def _style_dataframe_sheet(worksheet: Any, frame: pd.DataFrame) -> None:
    if not len(frame.columns):
        return
    header_fill = PatternFill("solid", fgColor="2563EB")
    header_font = Font(color="FFFFFF", bold=True)
    header_border = Border(bottom=Side(style="thin", color="B9C6D8"))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = header_border
    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.row_dimensions[1].height = 24
    for index, column in enumerate(frame.columns, start=1):
        samples = [str(column), *(str(value) for value in frame.iloc[:200, index - 1].fillna(""))]
        width = min(44, max(10, max((len(value) for value in samples), default=10) + 2))
        worksheet.column_dimensions[get_column_letter(index)].width = width


def _write_frame(
    frame: pd.DataFrame,
    options: OutputOptions,
    stem: str,
    sheet_name: str,
) -> Path | None:
    if len(frame) + 1 > EXCEL_MAX_ROWS or len(frame.columns) > EXCEL_MAX_COLUMNS:
        raise ValueError("Output exceeds the Excel worksheet row or column limit")
    if options.output_format == OutputFormat.CSV:
        output = _resolve_output(options.directory, stem, ".csv", options)
        if output is not None:
            frame.to_csv(output, index=False, encoding="utf-8-sig")
        return output
    output = _resolve_output(options.directory, stem, ".xlsx", options)
    if output is None:
        return None
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sanitize_sheet_name(sheet_name)
    for column_index, column in enumerate(frame.columns, start=1):
        worksheet.cell(1, column_index, str(column))
    for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
        for column_index, value in enumerate(row, start=1):
            worksheet.cell(row_index, column_index, _excel_safe_value(value))
    if options.style_output:
        _style_dataframe_sheet(worksheet, frame)
    workbook.save(output)
    workbook.close()
    return output


def _write_frames_workbook(
    frames: list[tuple[str, pd.DataFrame]],
    options: OutputOptions,
    stem: str,
    report_rows: list[dict[str, Any]] | None = None,
) -> Path | None:
    output = _resolve_output(options.directory, stem, ".xlsx", options)
    if output is None:
        return None
    workbook = Workbook()
    workbook.remove(workbook.active)
    used: set[str] = set()
    for preferred_name, frame in frames:
        name = _unique_sheet_name(preferred_name, used)
        worksheet = workbook.create_sheet(name)
        for column_index, column in enumerate(frame.columns, start=1):
            worksheet.cell(1, column_index, str(column))
        for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row_index, column_index, _excel_safe_value(value))
        if options.style_output:
            _style_dataframe_sheet(worksheet, frame)
    if options.create_report_sheet and report_rows:
        report = pd.DataFrame(report_rows)
        name = _unique_sheet_name("Task Report", used)
        worksheet = workbook.create_sheet(name)
        for column_index, column in enumerate(report.columns, start=1):
            worksheet.cell(1, column_index, str(column))
        for row_index, row in enumerate(report.itertuples(index=False, name=None), start=2):
            for column_index, value in enumerate(row, start=1):
                worksheet.cell(row_index, column_index, _excel_safe_value(value))
        if options.style_output:
            _style_dataframe_sheet(worksheet, report)
    if not workbook.sheetnames:
        workbook.create_sheet("Result")
    workbook.save(output)
    workbook.close()
    return output


def _copy_complete_sheet(source: Any, target_workbook: Workbook, title: str, preserve: bool) -> Any:
    target = target_workbook.create_sheet(
        _unique_sheet_name(title, set(target_workbook.sheetnames))
    )
    target.sheet_state = "visible"
    if preserve:
        for attribute in (
            "sheet_format",
            "sheet_properties",
            "page_margins",
            "page_setup",
            "print_options",
            "protection",
        ):
            with contextlib.suppress(Exception):
                setattr(target, attribute, copy.copy(getattr(source, attribute)))
        try:
            target.freeze_panes = source.freeze_panes
            target.sheet_view.showGridLines = source.sheet_view.showGridLines
            target.sheet_view.zoomScale = source.sheet_view.zoomScale
        except Exception:
            pass
    for row in source.iter_rows():
        for source_cell in row:
            target_cell = target.cell(source_cell.row, source_cell.column, source_cell.value)
            if not preserve:
                continue
            try:
                if source_cell.has_style:
                    target_cell._style = copy.copy(source_cell._style)
                if source_cell.hyperlink:
                    target_cell._hyperlink = copy.copy(source_cell.hyperlink)
                if source_cell.comment:
                    target_cell.comment = copy.copy(source_cell.comment)
            except Exception:
                pass
    if not preserve:
        return target
    for index, dimension in source.row_dimensions.items():
        target_dimension = target.row_dimensions[index]
        for attribute in ("height", "hidden", "outlineLevel", "collapsed"):
            with contextlib.suppress(Exception):
                setattr(target_dimension, attribute, getattr(dimension, attribute))
    for key, dimension in source.column_dimensions.items():
        target_dimension = target.column_dimensions[key]
        for attribute in ("width", "hidden", "bestFit", "outlineLevel", "collapsed", "min", "max"):
            with contextlib.suppress(Exception):
                setattr(target_dimension, attribute, getattr(dimension, attribute))
    for merged_range in source.merged_cells.ranges:
        with contextlib.suppress(Exception):
            target.merge_cells(str(merged_range))
    try:
        target.auto_filter.ref = source.auto_filter.ref
        target.print_area = str(source.print_area) if source.print_area else None
        target.print_title_rows = source.print_title_rows
        target.print_title_cols = source.print_title_cols
    except Exception:
        pass
    try:
        for validation in source.data_validations.dataValidation:
            target.add_data_validation(copy.deepcopy(validation))
    except Exception:
        pass
    try:
        for conditional_format, rules in source.conditional_formatting._cf_rules.items():
            for rule in rules:
                target.conditional_formatting.add(
                    str(conditional_format.sqref), copy.deepcopy(rule)
                )
    except Exception:
        pass
    try:
        for table in source.tables.values():
            target.add_table(copy.deepcopy(table))
    except Exception:
        pass
    for chart in getattr(source, "_charts", []):
        with contextlib.suppress(Exception):
            target.add_chart(copy.deepcopy(chart), copy.copy(chart.anchor))
    for image in getattr(source, "_images", []):
        with contextlib.suppress(Exception):
            target.add_image(copy.deepcopy(image), copy.copy(image.anchor))
    return target


def _unique_sheet_name(value: str, used: set[str]) -> str:
    base = sanitize_sheet_name(value)
    candidate = base
    index = 2
    while candidate.casefold() in {name.casefold() for name in used}:
        suffix = f"_{index}"
        candidate = f"{base[: 31 - len(suffix)]}{suffix}"
        index += 1
    used.add(candidate)
    return candidate


def _split_by_fields(
    frame: pd.DataFrame,
    fields: list[str],
    empty_label: str,
    blank_policy: BlankValuePolicy,
) -> list[tuple[str, pd.DataFrame]]:
    if not fields:
        raise ValueError("At least one split field is required")
    missing = [field for field in fields if field not in frame.columns]
    if missing:
        raise KeyError(f"Missing split fields: {', '.join(missing)}")
    grouping = frame[fields].copy()
    blank_mask = grouping.isna()
    for field in fields:
        blank_mask[field] |= grouping[field].astype(str).str.strip().eq("")
    if blank_policy == BlankValuePolicy.SKIP:
        keep = ~blank_mask.any(axis=1)
        frame = frame.loc[keep]
        grouping = grouping.loc[keep]
    grouping = grouping.where(~blank_mask.loc[grouping.index], empty_label)
    labels = grouping.astype(str).agg(
        lambda row: (
            " | ".join(f"{field}={value}" for field, value in zip(fields, row, strict=True))
            if len(fields) > 1
            else str(row.iloc[0])
        ),
        axis=1,
    )
    return [
        (label, frame.loc[labels == label].copy()) for label in labels.drop_duplicates().tolist()
    ]


def _fixed_groups(frame: pd.DataFrame, rows_per_file: int, balanced: bool) -> list[pd.DataFrame]:
    if rows_per_file <= 0:
        raise ValueError("Rows per output must be greater than zero")
    if not balanced:
        return [
            frame.iloc[start : start + rows_per_file].copy()
            for start in range(0, len(frame), rows_per_file)
        ] or [frame.copy()]
    count = max(1, (len(frame) + rows_per_file - 1) // rows_per_file)
    return _part_groups(frame, count)


def _part_groups(frame: pd.DataFrame, parts: int) -> list[pd.DataFrame]:
    if parts <= 0:
        raise ValueError("Part count must be greater than zero")
    quotient, remainder = divmod(len(frame), parts)
    groups: list[pd.DataFrame] = []
    start = 0
    for index in range(parts):
        size = quotient + (1 if index < remainder else 0)
        if size:
            groups.append(frame.iloc[start : start + size].copy())
        start += size
    return groups or [frame.copy()]


def _split_sheet_workbooks(
    source: Path,
    sheet_names: list[str],
    config: SplitTaskConfig,
    hooks: EngineHooks,
) -> list[Path]:
    outputs: list[Path] = []
    extension = source.suffix.casefold()
    workbook = (
        load_workbook(source, data_only=False, keep_vba=extension == ".xlsm")
        if extension in {".xlsx", ".xlsm"}
        else None
    )
    try:
        for index, sheet_name in enumerate(sheet_names, start=1):
            hooks.checkpoint()
            stem = render_naming_template(
                config.output.naming_template,
                {
                    "original_name": source.stem,
                    "sheet_name": sheet_name,
                    "split_field": "sheet",
                    "split_value": sheet_name,
                    "part_no": index,
                    "total_parts": len(sheet_names),
                    "index": index,
                },
            )
            stem = f"{config.output.prefix}{stem}{config.output.suffix}"
            output = _resolve_output(config.output.directory, stem, ".xlsx", config.output)
            if output is None:
                hooks.write_log(f"Skipped existing output: {stem}.xlsx")
                continue
            target = Workbook()
            target.remove(target.active)
            if workbook is not None and sheet_name in workbook.sheetnames:
                _copy_complete_sheet(
                    workbook[sheet_name],
                    target,
                    sheet_name,
                    config.output.preserve_format,
                )
            else:
                frame = read_sheet_frame(source, sheet_name, config.structure)
                worksheet = target.create_sheet(sanitize_sheet_name(sheet_name))
                for column_index, column in enumerate(frame.columns, start=1):
                    worksheet.cell(1, column_index, str(column))
                for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
                    for column_index, value in enumerate(row, start=1):
                        worksheet.cell(row_index, column_index, _excel_safe_value(value))
            target.save(output)
            target.close()
            outputs.append(output)
            hooks.write_log(f"Created: {output}")
    finally:
        if workbook is not None:
            workbook.close()
    return outputs


def execute_pro_split(
    config: SplitTaskConfig,
    hooks: EngineHooks | None = None,
) -> TaskResult:
    hooks = hooks or EngineHooks()
    result = TaskResult()
    input_rows = 0
    output_rows = 0
    total_sources = max(1, len(config.input_files))
    for source_index, path in enumerate(config.input_files, start=1):
        hooks.checkpoint()
        hooks.update(15 + int(55 * (source_index - 1) / total_sources), "Reading", path.name)
        try:
            sheet_names = _sheet_names(path, config.split.selected_sheets)
            if config.split.mode == SplitMode.BY_SHEET:
                outputs = _split_sheet_workbooks(path, sheet_names, config, hooks)
                result.output_files.extend(outputs)
                for sheet_name in sheet_names:
                    input_rows += len(read_sheet_frame(path, sheet_name, config.structure))
                output_rows += input_rows
                continue
            workbook_frames: list[tuple[str, pd.DataFrame]] = []
            report_rows: list[dict[str, Any]] = []
            for sheet_name in sheet_names:
                hooks.checkpoint()
                frame = read_sheet_frame(path, sheet_name, config.structure)
                input_rows += len(frame)
                if config.output.add_source_file:
                    frame["Source_File"] = path.name
                if config.output.add_source_sheet:
                    frame["Source_Sheet"] = sheet_name
                if config.split.mode == SplitMode.BY_FIELD:
                    groups: Iterable[tuple[str, pd.DataFrame]] = _split_by_fields(
                        frame,
                        config.split.fields,
                        config.split.empty_label,
                        config.split.blank_value_policy,
                    )
                    split_field = "-".join(config.split.fields)
                elif config.split.mode == SplitMode.FIXED_ROWS:
                    chunks = _fixed_groups(
                        frame,
                        config.split.rows_per_file,
                        config.split.distribution == SplitDistribution.BALANCED,
                    )
                    groups = [(str(index), chunk) for index, chunk in enumerate(chunks, 1)]
                    split_field = "rows"
                else:
                    chunks = _part_groups(frame, config.split.parts)
                    groups = [(str(index), chunk) for index, chunk in enumerate(chunks, 1)]
                    split_field = "parts"
                materialized = list(groups)
                for part_number, (split_value, group) in enumerate(materialized, start=1):
                    hooks.checkpoint()
                    output_rows += len(group)
                    stem = render_naming_template(
                        config.output.naming_template,
                        {
                            "original_name": path.stem,
                            "sheet_name": sheet_name,
                            "split_field": split_field,
                            "split_value": split_value,
                            "part_no": part_number,
                            "total_parts": len(materialized),
                            "row_count": len(group),
                            "index": source_index,
                        },
                    )
                    stem = f"{config.output.prefix}{stem}{config.output.suffix}"
                    report_rows.append(
                        {
                            "Source File": path.name,
                            "Source Sheet": sheet_name,
                            "Split Value": split_value,
                            "Rows": len(group),
                            "Output": stem,
                        }
                    )
                    if config.output.output_mode == OutputMode.SINGLE_WORKBOOK:
                        workbook_frames.append((stem, group))
                    else:
                        output = _write_frame(group, config.output, stem, stem)
                        if output is not None:
                            result.output_files.append(output)
                            hooks.write_log(f"Created: {output}")
                        else:
                            result.warnings.append(f"Skipped existing output: {stem}")
            if workbook_frames:
                workbook_stem = sanitize_filename(
                    f"{config.output.prefix}{config.output.workbook_name}{config.output.suffix}"
                )
                output = _write_frames_workbook(
                    workbook_frames,
                    config.output,
                    workbook_stem,
                    report_rows,
                )
                if output is not None:
                    result.output_files.append(output)
        except Exception as error:
            result.errors.append(f"{path.name}: {error}")
    result.reconciliation = reconcile_rows(input_rows, output_rows)
    if result.reconciliation.warning:
        result.warnings.append(result.reconciliation.warning)
    return result


def _ordered_union(column_sets: list[list[str]]) -> list[str]:
    return list(OrderedDict.fromkeys(column for columns in column_sets for column in columns))


def vertical_merge(frames: list[pd.DataFrame], strategy: FieldStrategy) -> pd.DataFrame:
    if not frames:
        return pd.DataFrame()
    columns = [list(frame.columns) for frame in frames]
    if strategy == FieldStrategy.STRICT:
        if any(current != columns[0] for current in columns[1:]):
            raise ValueError("Strict merge requires identical field names and order")
        target = columns[0]
    elif strategy == FieldStrategy.INTERSECTION:
        target = [column for column in columns[0] if all(column in other for other in columns[1:])]
    elif strategy == FieldStrategy.MASTER:
        target = columns[0]
    elif strategy == FieldStrategy.POSITION:
        target = columns[0]
        frames = [
            frame.set_axis(
                target[: len(frame.columns)]
                + [f"Column_{index}" for index in range(len(target) + 1, len(frame.columns) + 1)],
                axis=1,
            )
            for frame in frames
        ]
        target = _ordered_union([list(frame.columns) for frame in frames])
    else:
        target = _ordered_union(columns)
    return pd.concat([frame.reindex(columns=target) for frame in frames], ignore_index=True)


def _apply_dedupe(frame: pd.DataFrame, config: MergeTaskConfig) -> pd.DataFrame:
    mode = config.merge.dedupe_mode
    if mode == DedupeMode.NONE:
        return frame
    subset = None
    if mode == DedupeMode.FIELDS:
        missing = [field for field in config.merge.dedupe_fields if field not in frame.columns]
        if missing:
            raise KeyError(f"Missing deduplication fields: {', '.join(missing)}")
        subset = config.merge.dedupe_fields
    return frame.drop_duplicates(subset=subset, keep=config.merge.dedupe_keep).reset_index(
        drop=True
    )


def _aggregate(frame: pd.DataFrame, config: MergeTaskConfig) -> pd.DataFrame:
    groups = config.merge.aggregate_group_fields
    values = config.merge.aggregate_value_fields
    missing = [field for field in [*groups, *values] if field not in frame.columns]
    if missing:
        raise KeyError(f"Missing aggregation fields: {', '.join(missing)}")
    if not groups or not values:
        raise ValueError("Aggregation requires group fields and value fields")
    method = config.merge.aggregate_method
    if method == AggregateMethod.COUNT:
        return frame.groupby(groups, dropna=False)[values].count().reset_index()
    if method == AggregateMethod.NUNIQUE:
        return frame.groupby(groups, dropna=False)[values].nunique().reset_index()
    numeric = frame.copy()
    if method in {
        AggregateMethod.SUM,
        AggregateMethod.MEAN,
        AggregateMethod.MIN,
        AggregateMethod.MAX,
    }:
        for field in values:
            numeric[field] = pd.to_numeric(numeric[field], errors="coerce")
    return numeric.groupby(groups, dropna=False)[values].agg(method.value).reset_index()


def _copy_workbook_sources(
    config: MergeTaskConfig,
    hooks: EngineHooks,
) -> tuple[Path | None, int]:
    stem = render_naming_template(
        config.output.naming_template,
        {"original_name": "combined", "sheet_name": "workbook"},
    )
    output = _resolve_output(config.output.directory, stem, ".xlsx", config.output)
    if output is None:
        return None, 0
    target = Workbook()
    target.remove(target.active)
    copied_rows = 0
    for path in config.input_files:
        hooks.checkpoint()
        extension = path.suffix.casefold()
        if extension not in {".xlsx", ".xlsm"}:
            for sheet_name in _sheet_names(path, config.merge.selected_sheets):
                frame = read_sheet_frame(path, sheet_name, config.structure)
                copied_rows += len(frame)
                worksheet = target.create_sheet(
                    _unique_sheet_name(f"{path.stem}_{sheet_name}", set(target.sheetnames))
                )
                for column_index, column in enumerate(frame.columns, start=1):
                    worksheet.cell(1, column_index, str(column))
                for row_index, row in enumerate(frame.itertuples(index=False, name=None), start=2):
                    for column_index, value in enumerate(row, start=1):
                        worksheet.cell(row_index, column_index, _excel_safe_value(value))
                continue
        source = load_workbook(path, data_only=False, keep_vba=extension == ".xlsm")
        try:
            for sheet_name in _sheet_names(path, config.merge.selected_sheets):
                copied_rows += max(0, source[sheet_name].max_row - 1)
                preferred = f"{path.stem}_{sheet_name}"
                existing = next(
                    (name for name in target.sheetnames if name.casefold() == preferred.casefold()),
                    None,
                )
                if existing and config.merge.sheet_conflict_policy == SheetConflictPolicy.SKIP:
                    continue
                if existing and config.merge.sheet_conflict_policy == SheetConflictPolicy.REPLACE:
                    target.remove(target[existing])
                _copy_complete_sheet(
                    source[sheet_name],
                    target,
                    preferred,
                    config.merge.preserve_format or config.output.preserve_format,
                )
        finally:
            source.close()
    if not target.sheetnames:
        target.create_sheet("Result")
    target.save(output)
    target.close()
    return output, copied_rows


def execute_pro_merge(
    config: MergeTaskConfig,
    hooks: EngineHooks | None = None,
) -> TaskResult:
    hooks = hooks or EngineHooks()
    result = TaskResult()
    if config.merge.mode == MergeMode.WORKBOOK:
        try:
            output, rows = _copy_workbook_sources(config, hooks)
            if output is not None:
                result.output_files.append(output)
            result.reconciliation = ReconciliationResult(rows, rows)
        except Exception as error:
            result.errors.append(str(error))
        return result
    loaded: list[tuple[Path, str, pd.DataFrame]] = []
    input_rows = 0
    for path in config.input_files:
        hooks.checkpoint()
        try:
            for sheet_name in _sheet_names(path, config.merge.selected_sheets):
                if (
                    config.merge.mode == MergeMode.SAME_NAME
                    and config.merge.target_sheet_name
                    and sheet_name.casefold() != config.merge.target_sheet_name.casefold()
                ):
                    continue
                frame = read_sheet_frame(path, sheet_name, config.structure)
                input_rows += len(frame)
                if config.merge.add_source_file:
                    frame["Source_File"] = path.name
                if config.merge.add_source_sheet:
                    frame["Source_Sheet"] = sheet_name
                loaded.append((path, sheet_name, frame))
        except Exception as error:
            result.errors.append(f"{path.name}: {error}")
    output_rows = 0
    try:
        report_rows = [
            {"Source File": path.name, "Source Sheet": sheet, "Rows": len(frame)}
            for path, sheet, frame in loaded
        ]
        if config.merge.mode == MergeMode.SAME_NAME:
            groups: OrderedDict[str, tuple[str, list[pd.DataFrame]]] = OrderedDict()
            for _path, sheet_name, frame in loaded:
                key = sheet_name.casefold()
                groups.setdefault(key, (sheet_name, []))[1].append(frame)
            frames: list[tuple[str, pd.DataFrame]] = []
            for sheet_name, source_frames in groups.values():
                merged = _apply_dedupe(
                    vertical_merge(source_frames, config.merge.field_strategy), config
                )
                output_rows += len(merged)
                frames.append((f"{sheet_name}_Combined", merged))
            stem = render_naming_template(
                config.output.naming_template,
                {"original_name": "combined", "sheet_name": "same_name", "row_count": output_rows},
            )
            if config.output.output_format == OutputFormat.CSV and len(frames) == 1:
                output = _write_frame(frames[0][1], config.output, stem, frames[0][0])
            else:
                output = _write_frames_workbook(frames, config.output, stem, report_rows)
        else:
            merged = vertical_merge(
                [frame for _path, _sheet, frame in loaded],
                config.merge.field_strategy,
            )
            merged = _apply_dedupe(merged, config)
            if config.merge.mode == MergeMode.AGGREGATE:
                merged = _aggregate(merged, config)
            output_rows = len(merged)
            stem = render_naming_template(
                config.output.naming_template,
                {
                    "original_name": "combined",
                    "sheet_name": (
                        "Aggregate" if config.merge.mode == MergeMode.AGGREGATE else "Combined"
                    ),
                    "row_count": output_rows,
                },
            )
            output = _write_frame(
                merged,
                config.output,
                stem,
                "Aggregate" if config.merge.mode == MergeMode.AGGREGATE else "Combined",
            )
        if output is not None:
            result.output_files.append(output)
    except Exception as error:
        result.errors.append(str(error))
    result.reconciliation = reconcile_rows(input_rows, output_rows)
    if config.merge.mode != MergeMode.AGGREGATE and result.reconciliation.warning:
        result.warnings.append(result.reconciliation.warning)
    return result


def _normalize_key(value: Any, normalize: bool) -> str:
    if value is None or pd.isna(value):
        return ""
    text = str(value)
    if normalize:
        return unicodedata.normalize("NFKC", text).strip().casefold()
    return text


def _fuzzy_remap(
    main: pd.Series,
    related: pd.Series,
    threshold: int,
) -> pd.Series:
    main_values = list(OrderedDict.fromkeys(value for value in main.tolist() if value))
    cache: dict[str, str] = {}
    for value in OrderedDict.fromkeys(value for value in related.tolist() if value):
        best = ""
        best_score = 0.0
        for candidate in main_values:
            score = SequenceMatcher(None, value, candidate).ratio() * 100
            if score > best_score:
                best, best_score = candidate, score
        cache[value] = best if best_score >= threshold else value
    return related.map(lambda value: cache.get(value, value))


def _prepare_join_frame(
    frame: pd.DataFrame,
    keys: list[str],
    normalize: bool,
    empty_match: bool,
    prefix: str,
) -> tuple[pd.DataFrame, list[str]]:
    missing = [key for key in keys if key not in frame.columns]
    if missing:
        raise KeyError(f"Missing join keys: {', '.join(missing)}")
    prepared = frame.copy()
    join_keys: list[str] = []
    for position, key in enumerate(keys):
        internal = f"__join_{position}"
        values = prepared[key].map(lambda value: _normalize_key(value, normalize))
        if not empty_match:
            values = pd.Series(
                [
                    value if value else f"__empty_{prefix}_{index}"
                    for index, value in enumerate(values)
                ],
                index=prepared.index,
            )
        prepared[internal] = values
        join_keys.append(internal)
    return prepared, join_keys


def _dedupe_related(
    frame: pd.DataFrame,
    keys: list[str],
    policy: DuplicateKeyPolicy,
) -> tuple[pd.DataFrame, int]:
    duplicate_count = int(frame.duplicated(subset=keys, keep=False).sum())
    if not duplicate_count or policy == DuplicateKeyPolicy.EXPAND:
        return frame, duplicate_count
    if policy == DuplicateKeyPolicy.REJECT:
        raise ValueError(f"Duplicate related keys found: {duplicate_count} rows")
    if policy in {DuplicateKeyPolicy.FIRST, DuplicateKeyPolicy.LAST}:
        return frame.drop_duplicates(subset=keys, keep=policy.value), duplicate_count
    numeric = [
        column for column in frame.select_dtypes(include="number").columns if column not in keys
    ]
    aggregation = {
        column: ("sum" if column in numeric else "first")
        for column in frame.columns
        if column not in keys
    }
    return frame.groupby(keys, dropna=False, as_index=False).agg(aggregation), duplicate_count


def _coalesce_conflicts(
    merged: pd.DataFrame,
    overlap: list[str],
    policy: JoinConflictPolicy,
    main_suffix: str,
    related_suffix: str,
) -> pd.DataFrame:
    if policy == JoinConflictPolicy.KEEP_BOTH:
        return merged
    for column in overlap:
        main_column = f"{column}{main_suffix}"
        related_column = f"{column}{related_suffix}"
        if main_column not in merged.columns or related_column not in merged.columns:
            continue
        primary, fallback = (
            (main_column, related_column)
            if policy == JoinConflictPolicy.PREFER_MAIN
            else (related_column, main_column)
        )
        series = merged[primary]
        blank = series.isna() | series.astype(str).str.strip().eq("")
        merged[column] = series.where(~blank, merged[fallback])
        merged = merged.drop(columns=[main_column, related_column])
    return merged


def _key_join(
    main: pd.DataFrame,
    related: pd.DataFrame,
    config: AdvancedMergeTaskConfig,
) -> tuple[pd.DataFrame, dict[str, int]]:
    if not config.left_keys or len(config.left_keys) != len(config.right_keys):
        raise ValueError("Main and related key counts must be equal and non-zero")
    left, left_join_keys = _prepare_join_frame(
        main,
        config.left_keys,
        config.normalize_keys,
        config.empty_keys_match,
        "main",
    )
    right, right_join_keys = _prepare_join_frame(
        related,
        config.right_keys,
        config.normalize_keys,
        config.empty_keys_match,
        "related",
    )
    if config.fuzzy_match:
        if len(left_join_keys) != 1:
            raise ValueError("Fuzzy matching supports one key pair")
        right[right_join_keys[0]] = _fuzzy_remap(
            left[left_join_keys[0]],
            right[right_join_keys[0]],
            config.fuzzy_threshold,
        )
    right, duplicates = _dedupe_related(right, right_join_keys, config.duplicate_policy)
    for left_key, right_key in zip(left_join_keys, right_join_keys, strict=True):
        right[left_key] = right[right_key]
    right = right.drop(columns=right_join_keys)
    overlap = [
        column
        for column in main.columns.intersection(related.columns)
        if column not in set(config.left_keys) | set(config.right_keys)
    ]
    indicator = "__merge_status"
    merged = left.merge(
        right,
        on=left_join_keys,
        how=config.join_type.value,
        suffixes=(config.main_suffix, config.related_suffix),
        indicator=indicator,
    )
    status = merged[indicator].value_counts().to_dict()
    merged = merged.drop(columns=[*left_join_keys, indicator])
    merged = _coalesce_conflicts(
        merged,
        overlap,
        config.conflict_policy,
        config.main_suffix,
        config.related_suffix,
    )
    return merged, {
        "duplicates": duplicates,
        "matched": int(status.get("both", 0)),
        "main_only": int(status.get("left_only", 0)),
        "related_only": int(status.get("right_only", 0)),
    }


def _horizontal(
    frames: list[pd.DataFrame],
    labels: list[str],
    prefix: bool,
    gap: int,
) -> pd.DataFrame:
    if not frames:
        return pd.DataFrame()
    columns: list[pd.DataFrame] = []
    used: set[str] = set()
    for index, (label, frame) in enumerate(zip(labels, frames, strict=True), start=1):
        current = frame.reset_index(drop=True).copy()
        mapping: dict[str, str] = {}
        for column in current.columns:
            preferred = f"{label} · {column}" if prefix else str(column)
            candidate = preferred
            counter = 2
            while candidate in used:
                candidate = f"{preferred}_{counter}"
                counter += 1
            mapping[str(column)] = candidate
            used.add(candidate)
        columns.append(current.rename(columns=mapping))
        if gap and index < len(frames):
            columns.append(pd.DataFrame({f"__gap_{index}_{n}": "" for n in range(1, gap + 1)}))
    return pd.concat(columns, axis=1)


def execute_pro_advanced_merge(
    config: AdvancedMergeTaskConfig,
    hooks: EngineHooks | None = None,
) -> TaskResult:
    hooks = hooks or EngineHooks()
    result = TaskResult()
    frames: list[pd.DataFrame] = []
    labels: list[str] = []
    for path in config.input_files:
        hooks.checkpoint()
        try:
            sheet_name = _sheet_names(path, config.selected_sheets)[0]
            frame = read_sheet_frame(path, sheet_name, config.structure)
            frame = frame.rename(columns=config.field_mapping)
            if config.cleaning.trim_whitespace:
                frame = frame.map(lambda value: value.strip() if isinstance(value, str) else value)
            if config.cleaning.collapse_whitespace:
                frame = frame.map(
                    lambda value: re.sub(r"\s+", " ", value) if isinstance(value, str) else value
                )
            if config.cleaning.normalize_width:
                frame = frame.map(
                    lambda value: (
                        unicodedata.normalize("NFKC", value) if isinstance(value, str) else value
                    )
                )
            if config.cleaning.drop_empty_rows:
                frame = frame.replace(r"^\s*$", pd.NA, regex=True).dropna(how="all")
            if config.cleaning.drop_empty_columns:
                frame = frame.dropna(axis=1, how="all")
            if config.cleaning.drop_duplicate_rows:
                frame = frame.drop_duplicates()
            frames.append(frame.reset_index(drop=True))
            labels.append(f"{path.stem}_{sheet_name}")
        except Exception as error:
            result.errors.append(f"{path.name}: {error}")
    if len(frames) < 2:
        result.errors.append("Advanced merge requires at least two readable inputs")
        return result
    try:
        if config.mode == MergeMode.HORIZONTAL:
            merged = _horizontal(frames, labels, config.side_prefix, config.side_gap)
            input_reference = max(len(frame) for frame in frames)
        elif config.mode == MergeMode.KEY_JOIN:
            merged = frames[0]
            input_reference = len(merged)
            for related in frames[1:]:
                hooks.checkpoint()
                merged, stats = _key_join(merged, related, config)
                result.warnings.append(
                    "Join diagnostics: "
                    f"matched={stats['matched']}, main_only={stats['main_only']}, "
                    f"related_only={stats['related_only']}, duplicates={stats['duplicates']}"
                )
        else:
            raise ValueError(f"Unsupported advanced merge mode: {config.mode}")
        stem = render_naming_template(
            config.output.naming_template,
            {
                "original_name": "combined",
                "sheet_name": config.mode.value,
                "row_count": len(merged),
            },
        )
        output = _write_frame(
            merged,
            config.output,
            stem,
            "Horizontal" if config.mode == MergeMode.HORIZONTAL else "Joined",
        )
        if output is not None:
            result.output_files.append(output)
        result.reconciliation = ReconciliationResult(input_reference, len(merged))
    except Exception as error:
        result.errors.append(str(error))
    return result
