from __future__ import annotations

from collections.abc import Iterable, Sequence
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from excel_studio.models.task import HeaderMode, TableStructure


def make_unique_headers(values: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    counts: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        base = str(value).strip() if value not in (None, "") else f"Column_{index}"
        count = counts.get(base, 0) + 1
        counts[base] = count
        headers.append(base if count == 1 else f"{base}_{count}")
    return headers


def _headers_from_rows(rows: Sequence[Sequence[Any]], separator: str = "_") -> list[str]:
    width = max((len(row) for row in rows), default=0)
    combined: list[str] = []
    for column in range(width):
        parts = [
            str(row[column]).strip()
            for row in rows
            if column < len(row) and row[column] not in (None, "")
        ]
        combined.append(separator.join(parts) if parts else f"Column_{column + 1}")
    return make_unique_headers(combined)


def read_sheet_frame(path: Path, sheet_name: str, structure: TableStructure) -> pd.DataFrame:
    extension = path.suffix.casefold()
    if extension in {".csv", ".tsv"}:
        separator = "\t" if extension == ".tsv" else ","
        header = None if structure.header_mode == HeaderMode.NONE else structure.header_row - 1
        frame = pd.read_csv(path, sep=separator, header=header)
        if header is None:
            frame.columns = make_unique_headers(frame.columns)
        return _clean_frame(frame, structure)
    if extension not in {".xlsx", ".xlsm"}:
        header: int | list[int] | None
        if structure.header_mode == HeaderMode.NONE:
            header = None
        elif structure.header_mode == HeaderMode.MULTI_ROW:
            header = list(range(structure.header_row - 1, structure.header_end_row))
        else:
            header = structure.header_row - 1
        frame = pd.read_excel(path, sheet_name=sheet_name, header=header)
        if isinstance(frame.columns, pd.MultiIndex):
            frame.columns = make_unique_headers(
                "_".join(str(part) for part in column if not str(part).startswith("Unnamed"))
                for column in frame.columns
            )
        return _clean_frame(frame, structure)
    workbook = load_workbook(
        path,
        read_only=True,
        data_only=False,
        keep_vba=extension == ".xlsm",
        keep_links=False,
    )
    try:
        worksheet = workbook[sheet_name]
        start_column = max(1, structure.data_start_column)
        end_column = structure.data_end_column or worksheet.max_column
        rows = [
            list(row)
            for row in worksheet.iter_rows(
                min_row=1,
                max_row=structure.data_end_row or worksheet.max_row,
                min_col=start_column,
                max_col=end_column,
                values_only=True,
            )
        ]
    finally:
        workbook.close()
    if structure.skip_bottom_rows:
        rows = rows[: -structure.skip_bottom_rows]
    if structure.header_mode == HeaderMode.NONE:
        width = max((len(row) for row in rows), default=0)
        headers = [f"Column_{index + 1}" for index in range(width)]
        data_start = max(0, structure.data_start_row - 1)
    else:
        header_start = max(0, structure.header_row - 1)
        header_end = (
            max(header_start, structure.header_end_row - 1)
            if structure.header_mode == HeaderMode.MULTI_ROW
            else header_start
        )
        headers = _headers_from_rows(rows[header_start : header_end + 1])
        data_start = max(structure.data_start_row - 1, header_end + 1)
    data = rows[data_start:]
    normalized = [
        list(row[: len(headers)]) + [None] * max(0, len(headers) - len(row)) for row in data
    ]
    return _clean_frame(pd.DataFrame(normalized, columns=headers), structure)


def _clean_frame(frame: pd.DataFrame, structure: TableStructure) -> pd.DataFrame:
    cleaned = frame.copy()
    cleaned.columns = make_unique_headers(cleaned.columns)
    if structure.drop_empty_rows:
        cleaned = cleaned.dropna(axis=0, how="all")
    if structure.drop_empty_columns:
        cleaned = cleaned.dropna(axis=1, how="all")
    return cleaned.reset_index(drop=True)
