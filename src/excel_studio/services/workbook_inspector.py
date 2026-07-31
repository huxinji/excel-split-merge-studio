from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from excel_studio.models.workbook import FileRecord, PreviewResult, SheetInfo
from excel_studio.services.header_detector import detect_header


class WorkbookInspectionError(RuntimeError):
    pass


def _top_rows(worksheet: Any, limit: int = 20) -> list[list[Any]]:
    return [list(row) for row in worksheet.iter_rows(min_row=1, max_row=limit, values_only=True)]


def inspect_workbook(path: Path) -> list[SheetInfo]:
    extension = path.suffix.casefold()
    if extension in {".xlsx", ".xlsm"}:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=extension == ".xlsm",
            keep_links=False,
        )
        try:
            sheets: list[SheetInfo] = []
            for worksheet in workbook.worksheets:
                rows = _top_rows(worksheet)
                detection = detect_header(rows)
                is_empty = (
                    not any(value not in (None, "") for row in rows for value in row)
                    and worksheet.max_row <= 1
                )
                sheets.append(
                    SheetInfo(
                        name=worksheet.title,
                        state=worksheet.sheet_state,
                        max_row=worksheet.max_row,
                        max_column=worksheet.max_column,
                        is_empty=is_empty,
                        header_row=detection.row_number,
                        header_confidence=detection.confidence,
                    )
                )
            return sheets
        finally:
            workbook.close()
    if extension in {".csv", ".tsv"}:
        delimiter = "\t" if extension == ".tsv" else ","
        frame = pd.read_csv(path, sep=delimiter, nrows=20)
        return [
            SheetInfo(
                name=path.stem,
                max_row=len(frame) + 1,
                max_column=len(frame.columns),
                is_empty=frame.empty,
                header_row=1,
                header_confidence=1.0,
            )
        ]
    if extension in {".xls", ".xlsb"}:
        try:
            excel_file = pd.ExcelFile(path)
            return [SheetInfo(name=name, is_empty=False) for name in excel_file.sheet_names]
        except ImportError as error:
            raise WorkbookInspectionError(
                f"Optional reader for {extension} is not installed"
            ) from error
    raise WorkbookInspectionError(f"Unsupported file type: {extension}")


def build_file_record(path: Path) -> FileRecord:
    stat = path.stat()
    record = FileRecord(
        path=path,
        size_bytes=stat.st_size,
        modified_at=datetime.fromtimestamp(stat.st_mtime),
        extension=path.suffix.casefold(),
    )
    try:
        record.sheets = inspect_workbook(path)
        if not record.sheets:
            record.warning = "No worksheets found"
    except Exception as error:  # one damaged file must not abort a batch scan
        record.status = "error"
        record.warning = str(error)
    return record


def preview_workbook(
    path: Path,
    sheet_name: str | None = None,
    header_mode: str = "auto",
    header_row: int = 1,
    header_end_row: int | None = None,
    preview_rows: int = 20,
) -> PreviewResult:
    extension = path.suffix.casefold()
    if extension in {".csv", ".tsv"}:
        separator = "\t" if extension == ".tsv" else ","
        frame = pd.read_csv(path, sep=separator, nrows=preview_rows)
        header = detect_header([list(frame.columns), *frame.head(5).values.tolist()])
        return PreviewResult(
            path=path,
            sheet_name=path.stem,
            columns=[str(value) for value in frame.columns],
            rows=frame.fillna("").values.tolist(),
            total_rows=len(frame),
            total_columns=len(frame.columns),
            header=header,
        )
    if extension not in {".xlsx", ".xlsm"}:
        frame = pd.read_excel(
            path, sheet_name=sheet_name or 0, header=None, nrows=preview_rows + 20
        )
        raw_rows = frame.where(pd.notna(frame), None).values.tolist()
        selected_name = sheet_name or "Sheet1"
        max_row, max_column = frame.shape
    else:
        workbook = load_workbook(
            path,
            read_only=True,
            data_only=False,
            keep_vba=extension == ".xlsm",
            keep_links=False,
        )
        try:
            worksheet = workbook[sheet_name] if sheet_name else workbook.active
            selected_name = worksheet.title
            max_row, max_column = worksheet.max_row, worksheet.max_column
            limit = min(max_row, max(preview_rows + 20, 40))
            raw_rows = [
                list(row) for row in worksheet.iter_rows(min_row=1, max_row=limit, values_only=True)
            ]
        finally:
            workbook.close()
    detection = detect_header(raw_rows)
    if header_mode == "auto":
        start = detection.row_number - 1
        end = start
    elif header_mode == "none":
        start = -1
        end = -1
    elif header_mode == "multi_row":
        start = max(0, header_row - 1)
        end = max(start, (header_end_row or header_row) - 1)
        detection.row_number = header_row
    else:
        start = max(0, header_row - 1)
        end = start
        detection.row_number = header_row
    if start < 0:
        column_count = max((len(row) for row in raw_rows), default=0)
        columns = [f"Column_{index + 1}" for index in range(column_count)]
        data_rows = raw_rows[:preview_rows]
    else:
        header_rows = raw_rows[start : end + 1]
        column_count = max((len(row) for row in header_rows), default=0)
        columns = []
        for column_index in range(column_count):
            parts = [
                str(row[column_index]).strip()
                for row in header_rows
                if column_index < len(row) and row[column_index] not in (None, "")
            ]
            columns.append("_".join(parts) if parts else f"Column_{column_index + 1}")
        data_rows = raw_rows[end + 1 : end + 1 + preview_rows]
    normalized_rows = [
        list(row[: len(columns)]) + [None] * max(0, len(columns) - len(row)) for row in data_rows
    ]
    return PreviewResult(
        path=path,
        sheet_name=selected_name,
        columns=columns,
        rows=normalized_rows,
        total_rows=max(0, max_row - max(end + 1, 0)),
        total_columns=max_column,
        header=detection,
    )
