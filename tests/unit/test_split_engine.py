from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_studio.models.operations import (
    OutputOptions,
    SplitMode,
    SplitOptions,
    SplitTaskConfig,
)
from excel_studio.models.task import HeaderMode, TableStructure
from excel_studio.services.split_engine import (
    execute_split,
    fixed_row_groups,
    split_by_fields,
    split_into_parts,
)
from excel_studio.utils.naming import sanitize_filename, unique_output_path


class SplitEngineTests(unittest.TestCase):
    def test_2400_rows_fixed_limit_500(self) -> None:
        frame = pd.DataFrame({"ID": range(2400)})
        groups = fixed_row_groups(frame, 500)
        self.assertEqual([len(group) for group in groups], [500, 500, 500, 500, 400])

    def test_2400_rows_balanced_around_500(self) -> None:
        frame = pd.DataFrame({"ID": range(2400)})
        groups = fixed_row_groups(frame, 500, balanced=True)
        self.assertEqual([len(group) for group in groups], [480, 480, 480, 480, 480])

    def test_2400_rows_into_five_parts(self) -> None:
        frame = pd.DataFrame({"ID": range(2400)})
        groups = split_into_parts(frame, 5)
        self.assertEqual([len(group) for group in groups], [480, 480, 480, 480, 480])

    def test_split_by_multiple_fields_and_empty(self) -> None:
        frame = pd.DataFrame({"Area": ["A", "A", "B", None], "Branch": ["1", "2", "1", ""]})
        groups = split_by_fields(frame, ["Area", "Branch"])
        self.assertEqual([name for name, _ in groups], ["A__1", "A__2", "B__1", "EMPTY__EMPTY"])
        self.assertEqual(sum(len(group) for _, group in groups), 4)

    def test_execute_field_split_writes_and_reconciles(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            output = root / "output"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Details"
            sheet.append(["ID", "Branch", "Amount"])
            sheet.append([1, "North", 10])
            sheet.append([2, "South", 20])
            sheet.append([3, "North", 30])
            workbook.save(source)
            config = SplitTaskConfig(
                input_files=[source],
                structure=TableStructure(
                    header_mode=HeaderMode.ROW_NUMBER,
                    header_row=1,
                    data_start_row=2,
                ),
                split=SplitOptions(mode=SplitMode.BY_FIELD, fields=["Branch"]),
                output=OutputOptions(
                    directory=output,
                    naming_template="{original_name}_{split_value}",
                ),
            )
            result = execute_split(config)
            self.assertFalse(result.errors)
            self.assertEqual(len(result.output_files), 2)
            self.assertIsNotNone(result.reconciliation)
            self.assertTrue(result.reconciliation.is_balanced)
            row_counts = []
            for path in result.output_files:
                written = load_workbook(path, read_only=True)
                row_counts.append(written.active.max_row - 1)
                written.close()
            self.assertEqual(sorted(row_counts), [1, 2])

    def test_naming_sanitizes_and_auto_numbers(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            directory = Path(temporary)
            self.assertEqual(sanitize_filename("bad:folder*name."), "bad_folder_name")
            first = unique_output_path(directory, "result")
            first.touch()
            second = unique_output_path(directory, "result")
            self.assertEqual(second.name, "result_2.xlsx")


if __name__ == "__main__":
    unittest.main()
