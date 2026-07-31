from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook

from excel_studio.models.task import ScanOptions
from excel_studio.services.file_scanner import collect_input_files, scan_files
from excel_studio.services.workbook_inspector import preview_workbook


class ScannerAndPreviewTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.workbook_path = self.root / "source.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Details"
        sheet.append(["Business Report", None, None])
        sheet.append(["ID", "Branch", "Amount"])
        sheet.append([1, "North", 10])
        sheet.append([2, "South", 20])
        hidden = workbook.create_sheet("Hidden")
        hidden.sheet_state = "hidden"
        workbook.create_sheet("Empty")
        workbook.save(self.workbook_path)

    def tearDown(self) -> None:
        self.temporary.cleanup()

    def test_scan_reads_sheet_metadata(self) -> None:
        records = scan_files([self.workbook_path], ScanOptions())
        self.assertEqual(len(records), 1)
        self.assertEqual(
            [sheet.name for sheet in records[0].sheets], ["Details", "Hidden", "Empty"]
        )
        self.assertEqual(records[0].sheets[1].state, "hidden")

    def test_temp_files_and_subfolders_are_filtered(self) -> None:
        temp_book = self.root / "~$locked.xlsx"
        temp_book.write_bytes(b"temporary")
        nested = self.root / "nested"
        nested.mkdir()
        nested_book = nested / "nested.xlsx"
        nested_book.write_bytes(self.workbook_path.read_bytes())
        shallow = collect_input_files([self.root], ScanOptions())
        deep = collect_input_files([self.root], ScanOptions(scan_subfolders=True))
        self.assertEqual(shallow, [self.workbook_path])
        self.assertEqual(deep, [nested_book, self.workbook_path])

    def test_preview_honors_explicit_header_row(self) -> None:
        result = preview_workbook(
            self.workbook_path,
            "Details",
            header_mode="row_number",
            header_row=2,
            preview_rows=20,
        )
        self.assertEqual(result.columns, ["ID", "Branch", "Amount"])
        self.assertEqual(result.rows[0], [1, "North", 10])
        self.assertEqual(result.total_rows, 2)
