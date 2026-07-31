from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_studio.models.advanced import (
    AdvancedMergeTaskConfig,
    JoinConflictPolicy,
    JoinType,
)
from excel_studio.models.operations import (
    AggregateMethod,
    BlankValuePolicy,
    MergeMode,
    MergeOptions,
    MergeTaskConfig,
    OutputFormat,
    OutputMode,
    OutputOptions,
    SplitMode,
    SplitOptions,
    SplitTaskConfig,
)
from excel_studio.models.task import HeaderMode, TableStructure
from excel_studio.services.pro_engine import (
    execute_pro_advanced_merge,
    execute_pro_merge,
    execute_pro_split,
)


def _structure() -> TableStructure:
    return TableStructure(
        header_mode=HeaderMode.ROW_NUMBER,
        header_row=1,
        data_start_row=2,
    )


def _write_table(path: Path, rows: list[list[object]], title: str = "Data") -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(path)
    workbook.close()


class OriginalFeatureCompatibilityTests(unittest.TestCase):
    def test_split_csv_blank_skip_and_csv_output(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.csv"
            source.write_text("ID,Branch\n1,A\n2,\n3,B\n", encoding="utf-8")
            config = SplitTaskConfig(
                input_files=[source],
                structure=_structure(),
                split=SplitOptions(
                    mode=SplitMode.BY_FIELD,
                    fields=["Branch"],
                    blank_value_policy=BlankValuePolicy.SKIP,
                ),
                output=OutputOptions(
                    directory=root / "output",
                    naming_template="{split_value}",
                    add_source_file=False,
                    add_source_sheet=False,
                    output_format=OutputFormat.CSV,
                ),
            )
            result = execute_pro_split(config)
            self.assertFalse(result.errors)
            self.assertEqual(len(result.output_files), 2)
            self.assertTrue(all(path.suffix == ".csv" for path in result.output_files))
            self.assertEqual(sum(len(pd.read_csv(path)) for path in result.output_files), 2)
            self.assertEqual(result.reconciliation.input_rows, 3)
            self.assertEqual(result.reconciliation.output_rows, 2)

    def test_split_fixed_rows_to_one_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            _write_table(source, [["ID"], [1], [2], [3], [4], [5]])
            config = SplitTaskConfig(
                input_files=[source],
                structure=_structure(),
                split=SplitOptions(mode=SplitMode.FIXED_ROWS, rows_per_file=2),
                output=OutputOptions(
                    directory=root / "output",
                    naming_template="Part_{part_no}",
                    add_source_file=False,
                    add_source_sheet=False,
                    output_mode=OutputMode.SINGLE_WORKBOOK,
                    workbook_name="Split_Result",
                ),
            )
            result = execute_pro_split(config)
            self.assertFalse(result.errors)
            self.assertEqual(len(result.output_files), 1)
            workbook = load_workbook(result.output_files[0], read_only=True)
            self.assertEqual(len(workbook.sheetnames), 4)
            workbook.close()
            self.assertEqual(result.reconciliation.output_rows, 5)

    def test_sheet_split_preserves_common_formatting(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "formatted.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Report"
            sheet.append(["ID", "Name"])
            sheet.append([1, "Alpha"])
            sheet.column_dimensions["A"].width = 24
            sheet.freeze_panes = "A2"
            workbook.save(source)
            workbook.close()
            config = SplitTaskConfig(
                input_files=[source],
                structure=_structure(),
                split=SplitOptions(mode=SplitMode.BY_SHEET),
                output=OutputOptions(
                    directory=root / "output",
                    naming_template="{sheet_name}",
                    preserve_format=True,
                    add_source_file=False,
                    add_source_sheet=False,
                ),
            )
            result = execute_pro_split(config)
            self.assertFalse(result.errors)
            copied = load_workbook(result.output_files[0])
            self.assertEqual(copied["Report"].column_dimensions["A"].width, 24)
            self.assertEqual(copied["Report"].freeze_panes, "A2")
            copied.close()

    def test_merge_aggregate_matches_original_mode(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "one.xlsx"
            second = root / "two.xlsx"
            _write_table(first, [["Branch", "Amount"], ["A", 10], ["B", 5]])
            _write_table(second, [["Branch", "Amount"], ["A", 7], ["B", 8]])
            config = MergeTaskConfig(
                input_files=[first, second],
                structure=_structure(),
                merge=MergeOptions(
                    mode=MergeMode.AGGREGATE,
                    add_source_file=False,
                    add_source_sheet=False,
                    aggregate_group_fields=["Branch"],
                    aggregate_value_fields=["Amount"],
                    aggregate_method=AggregateMethod.SUM,
                ),
                output=OutputOptions(
                    directory=root / "output",
                    naming_template="aggregate",
                    add_source_file=False,
                    add_source_sheet=False,
                ),
            )
            result = execute_pro_merge(config)
            self.assertFalse(result.errors)
            frame = pd.read_excel(result.output_files[0])
            values = dict(zip(frame["Branch"], frame["Amount"], strict=True))
            self.assertEqual(values, {"A": 17, "B": 13})

    def test_fuzzy_join_and_prefer_related_conflict(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            main = root / "main.xlsx"
            related = root / "related.xlsx"
            _write_table(main, [["Key", "Status"], ["North", ""]])
            _write_table(related, [["Key", "Status"], ["Nroth", "Approved"]])
            config = AdvancedMergeTaskConfig(
                input_files=[main, related],
                structure=_structure(),
                mode=MergeMode.KEY_JOIN,
                output=OutputOptions(
                    directory=root / "output",
                    naming_template="joined",
                    add_source_file=False,
                    add_source_sheet=False,
                ),
                join_type=JoinType.LEFT,
                left_keys=["Key"],
                right_keys=["Key"],
                fuzzy_match=True,
                fuzzy_threshold=50,
                conflict_policy=JoinConflictPolicy.PREFER_RELATED,
            )
            result = execute_pro_advanced_merge(config)
            self.assertFalse(result.errors)
            frame = pd.read_excel(result.output_files[0])
            self.assertEqual(frame.loc[0, "Status"], "Approved")

    def test_horizontal_gap_is_materialized(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "one.xlsx"
            second = root / "two.xlsx"
            _write_table(first, [["A"], [1], [2]])
            _write_table(second, [["B"], [3]])
            config = AdvancedMergeTaskConfig(
                input_files=[first, second],
                structure=_structure(),
                mode=MergeMode.HORIZONTAL,
                output=OutputOptions(directory=root / "output", naming_template="horizontal"),
                side_prefix=False,
                side_gap=2,
            )
            result = execute_pro_advanced_merge(config)
            self.assertFalse(result.errors)
            frame = pd.read_excel(result.output_files[0])
            self.assertEqual(len(frame), 2)
            self.assertEqual(len(frame.columns), 4)


if __name__ == "__main__":
    unittest.main()
