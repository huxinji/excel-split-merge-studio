from __future__ import annotations

import json
import tempfile
import threading
import time
import unittest
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook

from excel_studio.models.execution import TaskExecutionSummary
from excel_studio.models.operations import OutputOptions, SplitMode, SplitOptions, SplitTaskConfig
from excel_studio.models.result import TaskResult
from excel_studio.models.task import TableStructure
from excel_studio.services.history_service import HistoryStore
from excel_studio.services.report_service import write_task_reports
from excel_studio.workers.task_worker import TaskCancelled, TaskControl


class TaskSystemTests(unittest.TestCase):
    def test_pause_resume_and_cancel_are_thread_safe(self) -> None:
        control = TaskControl()
        control.pause()
        passed_checkpoint = threading.Event()

        def wait_at_checkpoint() -> None:
            control.checkpoint()
            passed_checkpoint.set()

        thread = threading.Thread(target=wait_at_checkpoint)
        thread.start()
        time.sleep(0.05)
        self.assertFalse(passed_checkpoint.is_set())
        control.resume()
        thread.join(timeout=1)
        self.assertTrue(passed_checkpoint.is_set())
        control.cancel()
        with self.assertRaises(TaskCancelled):
            control.checkpoint()

    def test_report_and_history_include_required_artifacts(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            Workbook().save(source)
            config = SplitTaskConfig(
                input_files=[source],
                structure=TableStructure(),
                split=SplitOptions(mode=SplitMode.FIXED_ROWS),
                output=OutputOptions(directory=root / "output"),
            )
            now = datetime.now()
            summary = TaskExecutionSummary(
                task_id="task-1",
                operation="split",
                status="success",
                started_at=now,
                finished_at=now,
                output_directory=root / "output",
                result=TaskResult(output_files=[root / "output" / "result.xlsx"]),
                log_messages=["started", "finished"],
            )
            report_directory = write_task_reports(summary, config)
            self.assertTrue((report_directory / "task-report.json").exists())
            self.assertTrue((report_directory / "task-report.xlsx").exists())
            self.assertTrue((report_directory / "task.log").exists())
            payload = json.loads(
                (report_directory / "task-report.json").read_text(encoding="utf-8")
            )
            self.assertEqual(payload["task_id"], "task-1")
            report_book = load_workbook(report_directory / "task-report.xlsx", read_only=True)
            self.assertEqual(
                report_book.sheetnames,
                [
                    "Summary",
                    "Input Files",
                    "Output Files",
                    "Warnings",
                    "Errors",
                    "Field Mapping",
                    "Row Reconciliation",
                ],
            )
            report_book.close()
            summary.report_directory = report_directory
            history = HistoryStore(root / "config")
            history.add(summary)
            self.assertEqual(history.load()[0]["task_id"], "task-1")


if __name__ == "__main__":
    unittest.main()
