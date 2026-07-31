from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

import pandas as pd
from openpyxl import Workbook, load_workbook

from excel_studio.models.operations import (
    FieldStrategy,
    MergeMode,
    MergeOptions,
    MergeTaskConfig,
    OutputOptions,
)
from excel_studio.models.task import HeaderMode, TableStructure
from excel_studio.services.merge_engine import execute_merge, vertical_merge


class MergeEngineTests(unittest.TestCase):
    def test_strict_merge_rejects_different_order(self) -> None:
        first = pd.DataFrame({"A": [1], "B": [2]})
        second = pd.DataFrame({"B": [3], "A": [4]})
        with self.assertRaises(ValueError):
            vertical_merge([first, second], FieldStrategy.STRICT)

    def test_union_aligns_different_fields_and_order(self) -> None:
        first = pd.DataFrame({"A": [1], "B": [2]})
        second = pd.DataFrame({"B": [3], "C": [4]})
        merged = vertical_merge([first, second], FieldStrategy.UNION)
        self.assertEqual(list(merged.columns), ["A", "B", "C"])
        self.assertEqual(len(merged), 2)
        self.assertTrue(pd.isna(merged.loc[1, "A"]))

    def test_intersection_keeps_common_fields(self) -> None:
        first = pd.DataFrame({"A": [1], "B": [2]})
        second = pd.DataFrame({"B": [3], "C": [4]})
        merged = vertical_merge([first, second], FieldStrategy.INTERSECTION)
        self.assertEqual(list(merged.columns), ["B"])

    def test_same_name_merge_writes_sources_and_reconciles(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            first = root / "first.xlsx"
            second = root / "second.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Details"
            sheet.append(["ID", "Amount"])
            sheet.append([1, 10])
            sheet.append([2, 20])
            workbook.save(first)
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Details"
            sheet.append(["Amount", "ID", "Extra"])
            sheet.append([30, 3, "x"])
            workbook.save(second)
            config = MergeTaskConfig(
                input_files=[first, second],
                structure=TableStructure(
                    header_mode=HeaderMode.ROW_NUMBER, header_row=1, data_start_row=2
                ),
                merge=MergeOptions(
                    mode=MergeMode.SAME_NAME,
                    field_strategy=FieldStrategy.UNION,
                    target_sheet_name="Details",
                ),
                output=OutputOptions(
                    directory=root / "output", naming_template="{sheet_name}_Combined"
                ),
            )
            result = execute_merge(config)
            self.assertFalse(result.errors)
            self.assertEqual(len(result.output_files), 1)
            self.assertTrue(result.reconciliation.is_balanced)
            frame = pd.read_excel(result.output_files[0])
            self.assertEqual(len(frame), 3)
            self.assertIn("Source_File", frame.columns)
            self.assertIn("Source_Sheet", frame.columns)
            self.assertIn("Extra", frame.columns)

    def test_workbook_level_merge_keeps_separate_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            source = root / "source.xlsx"
            workbook = Workbook()
            first = workbook.active
            first.title = "Summary"
            first.append(["ID"])
            first.append([1])
            second = workbook.create_sheet("Details")
            second.append(["ID"])
            second.append([2])
            workbook.save(source)
            result = execute_merge(
                MergeTaskConfig(
                    input_files=[source],
                    structure=TableStructure(
                        header_mode=HeaderMode.ROW_NUMBER, header_row=1, data_start_row=2
                    ),
                    merge=MergeOptions(mode=MergeMode.WORKBOOK),
                    output=OutputOptions(
                        directory=root / "output", naming_template="combined_workbook"
                    ),
                )
            )
            self.assertFalse(result.errors)
            merged = load_workbook(result.output_files[0], read_only=True)
            self.assertEqual(merged.sheetnames, ["source_Summary", "source_Details"])
            merged.close()


if __name__ == "__main__":
    unittest.main()
