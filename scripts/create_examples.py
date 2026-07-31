from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook


def create_examples(output_directory: Path) -> None:
    output_directory.mkdir(parents=True, exist_ok=True)
    sales = Workbook()
    details = sales.active
    details.title = "Details"
    details.append(["ID", "Branch", "Area", "Date", "Amount"])
    details.append([1, "North-01", "North", "2026-01-01", 1200])
    details.append([2, "South-01", "South", "2026-01-02", 800])
    details.append([3, "North-02", "North", "2026-02-01", 1500])
    summary = sales.create_sheet("Summary")
    summary.append(["Area", "Total"])
    summary.append(["North", 2700])
    summary.append(["South", 800])
    sales.save(output_directory / "sales-example.xlsx")

    related = Workbook()
    branches = related.active
    branches.title = "Branches"
    branches.append(["BranchCode", "Manager", "Region"])
    branches.append(["North-01", "Alice", "N"])
    branches.append(["North-02", "Bob", "N"])
    branches.append(["West-01", "Carol", "W"])
    related.save(output_directory / "branch-master-example.xlsx")


if __name__ == "__main__":
    create_examples(Path(__file__).resolve().parents[1] / "examples")
